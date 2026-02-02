import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import swifter
import openai
from dotenv import load_dotenv, find_dotenv
import os
from joblib import Memory
from rapidfuzz import process, fuzz 
from stqdm import stqdm
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm
import time
import requests
import os
import boto3
from smart_open import open 
from datetime import datetime 
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import s3fs

#change path where the code lies
script_dir = os.path.dirname(os.path.realpath('scs_app'))
os.chdir(script_dir)

print("Current working directory:", os.getcwd())


print("Current time:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
# pd.set_option('display.max_colwidth', None)

# st.set_page_config(
#     page_title="FTL RFQ Pricing Tool", # Changed page title here
#     page_icon="🚚", # Truck emoji for the browser tab icon
#     layout="wide"
# )

st.markdown(
    "<h1><span style='color:red;'>Delhivery</span> FTL RFQ Pricing</h1>",
    unsafe_allow_html=True
)
env_path = find_dotenv()
print("Loading .env from :", env_path)
load_dotenv(env_path)

AZURE_API_KEY = os.getenv("AZURE_API_KEY")
AZURE_ENDPOINT = os.getenv("AZURE_ENDPOINT")
DEPLOYMENT_NAME = os.getenv("DEPLOYMENT_NAME")
API_VERSION = os.getenv("API_VERSION")

google_distance_api_key = os.getenv("Google_distance_api_key")
locate_one_api_key = os.getenv("locate_one_api_key")

client = openai.AzureOpenAI(
    api_key=AZURE_API_KEY,
    api_version=API_VERSION,
    azure_endpoint=AZURE_ENDPOINT
)

def get_llm_suggestions(column_names):
    """Ask Azure LLM to infer key logistics column names and return structured JSON."""
    prompt = ( 
        f'''Given these column names: {column_names}, identify columns related to 
        origin, origin_state, destination, destination_state, vehicle_type, load_mt. 
        Respond in strict JSON format:
        {{'origin': 'col_name', 'origin_state':'col_name','destination': 'col_name', 'destination_state':'col_name',
        'vehicle_type': 'col_name', 'load_mt': 'col_name'}} 
        If a column is missing, return null (not a string, but actual JSON null). No extra text, just JSON.'''
    )

    #st.write("🔹 **Prompt Sent to Azure:**", prompt)  # ✅ Debugging print

    response = client.chat.completions.create(
        model=DEPLOYMENT_NAME,
        messages=[{"role": "system", "content": "You are an AI assistant."}, {"role": "user", "content": prompt}],
        max_tokens=300
    )

    raw_llm_output = response.choices[0].message.content
    # st.write("🔹 **Raw Azure Response:**", raw_llm_output)  # ✅ Debugging print
    cleaned_output = re.sub(r"```(?:json)?\s*([\s\S]*?)\s*```", r"\1", raw_llm_output)

    try:
        parsed_response = json.loads(cleaned_output)
        # st.write("🔹 **Parsed Azure Response:**", parsed_response)  # ✅ Debugging print
        return parsed_response
    except json.JSONDecodeError:
        st.write("⚠️ **Error: Unable to parse LLM response**")
        return {}


if 'rfq_odvtmt_output_df_session' not in st.session_state:
    st.session_state.rfq_odvtmt_output_df_session = None

if 'rfq_odvt_output_df_session' not in st.session_state:
    st.session_state.rfq_odvt_output_df_session = None

if 'summary_stats' not in st.session_state:
    st.session_state.summary_stats = None



def read_csv_s3():
    # Set up
    bucket = 'abis3'
    prefix = 'Sheetal/FTL_RFQ_Pricing_Data/FTL_RFQ_Pricing_query_data/'

    # List all CSV part files in the folder
    s3 = boto3.client('s3')
    objects = s3.list_objects_v2(Bucket=bucket, Prefix=prefix)

    # Filter only part CSV files
    csv_keys = [obj['Key'] for obj in objects.get('Contents', []) if obj['Key'].endswith('.csv')]

    # Read and concat all parts
    dfs = []
    for key in csv_keys:
        s3_path = f's3://{bucket}/{key}'
        with open(s3_path, 'r') as f:
            df = pd.read_csv(f)
           
            dfs.append(df)

    # Combine into a single DataFrame
    full_df = pd.concat(dfs, ignore_index=True)

    return full_df


def read_csv_s3_multiple():

    # source1 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/ODVT_FTL_Pricing_query_data/'}
    # source2 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/OCDVT_FTL_Pricing_query_data/'}
    # source3 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/OCDCVT_FTL_RFQ_Pricing_query_data/'}
    # source4 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/City_Cluster_mapping/'}

    source1 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/ODVT_FTL_Pricing_query_data_test/'}
    source2 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/OCDVT_FTL_Pricing_query_data_test/'}
    source3 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/OCDCVT_FTL_RFQ_Pricing_query_data_test/'}
    source4 = {'bucket': 'abis3', 'prefix': 'Sheetal/FTL_RFQ_Pricing_Data/City_Cluster_mapping/'}
    s3 = boto3.client('s3')

    def read_from_s3(source):
        dfs = []
        response = s3.list_objects_v2(Bucket=source['bucket'], Prefix=source['prefix'])
        csv_keys = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.csv')]

        for key in csv_keys:
            s3_path = f"s3://{source['bucket']}/{key}"
            with open(s3_path, 'r') as f:
                df = pd.read_csv(s3_path, low_memory=False)
                dfs.append(df)

        combined_df = pd.concat(dfs, ignore_index=True)
        return combined_df
    
    odvt_ftl_pricing_data = read_from_s3(source1)
    ocdvt_ftl_pricing_data = read_from_s3(source2)
    ocdcvt_ftl_rfq_pricing_data = read_from_s3(source3)
    city_cluster_df = read_from_s3(source4)

    return odvt_ftl_pricing_data, ocdvt_ftl_pricing_data, ocdcvt_ftl_rfq_pricing_data, city_cluster_df


# @st.cache_data(ttl=3600)  # Cache expires after 1 hour
@st.cache_data() 


def load_data(refresh_time=None):
    # Loading RFQ FTL pricing data 1 year & city-cluster mapping file 
    if os.getcwd() == '/Users/sheetal.suwalka/Documents/Python_codes/Missing_dimensions':
        RFQ_FTL_Pricing = pd.read_csv('/Users/sheetal.suwalka/Documents/Python_codes/Missing_dimensions/RFQ_FTL_pricing_tool/FTL_and_RFQ_Pricing.csv')
        city_cluster_mapping = pd.read_csv('/Users/sheetal.suwalka/Documents/Python_codes/Missing_dimensions/RFQ_FTL_pricing_tool/city_cluster_mapping.csv')
    else:
        # RFQ_FTL_Pricing = read_csv_s3() 
        ODVTFTL_pricing, OCDVTFTL_pricing, RFQ_FTL_Pricing, city_cluster_mapping = read_csv_s3_multiple()
        # city_cluster_mapping = pd.read_csv('city_cluster_mapping.csv')

    city_cluster_mapping = city_cluster_mapping[['city','state','city_code','cluster','cluster_new']]
    city_cluster_mapping['city'] = city_cluster_mapping['city'].str.lower()
    city_cluster_mapping['state'] = city_cluster_mapping['state'].str.upper()
    city_cluster_mapping['cluster'] = city_cluster_mapping['cluster'].str.lower()
    city_cluster_mapping = city_cluster_mapping[city_cluster_mapping['cluster'].notna()].drop_duplicates()

    return city_cluster_mapping, ODVTFTL_pricing, OCDVTFTL_pricing, RFQ_FTL_Pricing


###### Pincode Mapping Functions  - Starts #####
def fuzzy_match_on_pincode(df):
    # Check if required columns exist
    required_cols = {"mapped_pincode", "mapped_city_db"}
    if not required_cols.issubset(df.columns):
        print("Missing required columns. Skipping fuzzy match on pincode.")
        return df  # Return unchanged dataframe if columns are missing

    df["mapped_pincode"] = df["mapped_pincode"].astype(str).str.zfill(6)  # Ensure pincode is 6 digits
    df["pincode_prefix"] = df["mapped_pincode"].str[:3]  # Extract first 3 digits

    # Filter only rows where mapped_city_tat is missing
    missing_tat_df = df[df["mapped_city_tat"].isna()].copy()

    # If no rows need filling, return early
    if missing_tat_df.empty:
        return df

    # Create a lookup dictionary for pincode-prefix based matching
    lookup_df = df.dropna(subset=["mapped_city_tat"]).copy()
    lookup_df["mapped_pincode"] = lookup_df["mapped_pincode"].astype(int)  # Convert to int for numeric comparison
    lookup_dict = {}

    for (prefix, mapped_city_db), group in lookup_df.groupby(["pincode_prefix", "mapped_city_db"]):
        # Find the closest pincode numerically
        closest_match = group.loc[(group["mapped_pincode"] - group["mapped_pincode"].min()).idxmin()]
        lookup_dict[(prefix, mapped_city_db)] = {
            "mapped_city_tat": closest_match["mapped_city_tat"],
            "mapped_state_db": closest_match["mapped_state_db"]
        }

    # Apply fuzzy matching to fill missing mapped_city_tat
    def apply_fuzzy_match(row):
        key = (row["pincode_prefix"], row["mapped_city_db"])
        if key in lookup_dict:
            match = lookup_dict[key]
            row["mapped_city_tat"] = match["mapped_city_tat"]
            row["mapped_state_db"] = match["mapped_state_db"]
            row["mapping_type"] = "fuzzy match on pincode"
        return row

    df.update(missing_tat_df.apply(apply_fuzzy_match, axis=1))
    df.drop(columns=["pincode_prefix"], inplace=True, errors="ignore")  # Remove helper column
    return df

def validate_pincode(pincode):
    """Check if a pincode is a valid 6-digit Indian pincode."""
    return bool(re.fullmatch(r"[1-9][0-9]{5}", str(pincode)))

def read_pincode_csv_s3():
    # Set up
    bucket = 'abis3'
    prefix = 'Sheetal/pincode_city_mapping/Pincode_city_recent_mapping_with_Lat_Long'

    # List all CSV part files in the folder
    s3 = boto3.client('s3')
    objects = s3.list_objects_v2(Bucket=bucket, Prefix=prefix)

    # Filter only part CSV files
    csv_keys = [obj['Key'] for obj in objects.get('Contents', []) if obj['Key'].endswith('.csv')]

    # Read and concat all parts
    dfs = []
    for key in csv_keys:
        s3_path = f's3://{bucket}/{key}'
        with open(s3_path, 'r') as f:
            df = pd.read_csv(f)
            dfs.append(df)
    # Combine into a single DataFrame
    full_df = pd.concat(dfs, ignore_index=True)
    return full_df

def get_internal_pincode_db():
    """Fetch the internal pincode database from a local CSV file."""
    try:
        # df_pincode = pd.read_csv("Pincode_Mapping.csv")
        if  '/Users/sheetal.suwalka/Documents/Python_codes' in os.getcwd():
            df_pincode = pd.read_csv('/Users/sheetal.suwalka/Documents/Python_codes/Missing_dimensions/Pincode_Mapping.csv')
        else:
            df_pincode  = read_pincode_csv_s3()

        print('pincode mapping data shape:', df_pincode.shape)
        print('pincode data last records',df_pincode.tail())

        df_pincode.columns = df_pincode.columns.str.strip()
        df_pincode = df_pincode[df_pincode['mapped_pincode'].notna()]
        df_pincode["mapped_pincode"] = df_pincode["mapped_pincode"].astype(int).astype(str)
        df_pincode["Latitude"] = df_pincode["Latitude"].astype(str)
        df_pincode["Longitude"] = df_pincode["Longitude"].astype(str)
        df_pincode["mapped_city_db"] = df_pincode["mapped_city_db"].str.lower().fillna("Unknown")
        df_pincode["mapped_state_db"].fillna("Unknown", inplace=True)
        return df_pincode
    except FileNotFoundError:
        st.error("Error: Pincode_Mapping.csv file not found.")
        return pd.DataFrame()
    
def get_most_probable_pincode(row):
    """Use Azure LLM to get the most probable pincode for a given Indian city."""
    city = row['destination']
    state = row['destination_state'] if 'destination_state' in row and pd.notna(row['destination_state']) else None

    if state:
        state = row['destination_state']
        prompt = (
            f"Given the city '{city}' of the {state} state in India, return the most commonly used 6-digit Indian pincode for this city. "
            "The response MUST follow strict JSON format with double quotes: "
            "{\"mapped_pincode\": \"valid_pincode\"}. "
            "You MUST always return a valid 6-digit pincode—never leave it empty or return 'Unknown'."
        )
    else:
        prompt = (
            f"Given the city '{city}' in India, return the most commonly used 6-digit Indian pincode for this city. "
            "The response MUST follow strict JSON format with double quotes: "
            "{\"mapped_pincode\": \"valid_pincode\"}. "
            "You MUST always return a valid 6-digit pincode—never leave it empty or return 'Unknown'."
        )

    
    #st.write(f"🔹 Sending to Azure for missing pincode: {prompt}")  # ✅ Debugging print

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100
        )

        response_text = response.choices[0].message.content.strip()
        #st.write(f"🔹 Azure Response: {response_text}")  # ✅ Debugging print

        response_text = response_text.replace("```json", "").replace("```", "").strip()
        parsed_response = json.loads(response_text) if response_text.startswith("{") else {}

        probable_pincode = parsed_response.get("mapped_pincode", "000000")  # Default fallback

        return probable_pincode if re.fullmatch(r"[1-9][0-9]{5}", probable_pincode) else "000000"  # Validate
    except (json.JSONDecodeError, openai.BadRequestError):
        return "000000"  # Fallback pincode if error occurs
    
def get_nearest_major_city(row):
    """Use Azure LLM to get the nearest major city for a given Indian city."""
    city = row['city']

    state = row['state'] if 'state' in row and pd.notna(row['state']) else None

    if state:
        state = row['state']
        # prompt = (f''' 
        #         You are mapping Indian cities. Input will be a lesser-known Indian city. Output the nearest *well-known major Indian city*.
        #     For the given  city '{city}' of '{state}' state in India
        #     Rules:
        #     - Always return a valid known Indian city name
        #     - If multiple candidates exist, choose the closest by geography
        #     - If uncertain, pick the most plausible major city—never say "Unknown"

        #         Respond in strict JSON format 
        #         ```json
        #         {{'mapped_city': 'nearest_major_city_name'}}```
        #         You MUST always return a valid city name—never leave it empty or return 'Unknown'.'''
        #     )
        
        prompt = f"""
You are mapping Indian cities.

Input:
- A lesser-known Indian city name: '{city}'
- Its state: '{state}'

Task:
Map this city to the **nearest well-known major Indian city**.

Strict Rules:
1. The mapped city **MUST be within the same Indian state ('{state}')**.
2. The mapped city **MUST be within 100 km (road or aerial distance)** from the input city.
3. Choose a **well-known / major city** (district HQ, large urban center, or widely recognized city).
4. If multiple valid cities exist, select the **closest by geography**.
5. **DO NOT** return cities from another state unless **no valid city exists within 100 km in the same state**.
6. Never return placeholders like "Unknown", "N/A", or empty values.

Fallback Rule (only if unavoidable):
- If no major city exists within 100 km in the same state, return the **nearest major city in the same state even if slightly beyond 100 km**.
- Crossing state boundaries is a **last resort** and should be avoided unless unavoidable.

Response Format:
Respond in **strict JSON only**, with no extra text.

```json
    {{'mapped_city': 'nearest_major_city_name'}}```

You MUST always return a valid Indian city name.
"""


    else:
        prompt = (
                f''' 
                You are mapping Indian cities. Input will be a lesser-known Indian city. Output the nearest *well-known major Indian city*.
            For the given  city '{city}' in India
            Rules:
            - Always return a valid known Indian city name
            - If multiple candidates exist, choose the closest by geography
            - If uncertain, pick the most plausible major city—never say "Unknown"

                Respond in strict JSON format 
                ```json
                {{'mapped_city': 'nearest_major_city_name'}}```
                You MUST always return a valid city name—never leave it empty or return 'Unknown'.'''
            )

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100,
            temperature=0  # Set temperature to 0 for deterministic output
        )

        response_text =  response.choices[0].message.content

        try:
            json_match = re.search(r"```json\n([\s\S]+?)\n```", response_text)
            if json_match:
                json_str = json_match.group(1) 
                json_str = json_str.replace("'", '"')
                json_data = json.loads(json_str)
                return pd.Series([city,state,  json_data.get('mapped_city')])
            
            else:
                pd.Series([city,state, None])

        except:
            return pd.Series([city,state, None])
        
    except:
            return pd.Series([city,state, None])


def get_correct_pincode_or_city(pincode):
    """Use Azure LLM to correct pincode and find mapped city."""
    
    prompt = (
        f"The given Indian pincode is '{pincode}', please correct it."
        "If this pincode is too long (more than 6 digits), use only the first 6 digits. "
        "If the pincode is too short (less than 6 digits), try adding a suitable last digit to make it valid. "
        "If the last digit seems incorrect, try changing it to the closest valid number. "
        "Always return the closest correct 6-digit pincode based on known patterns. "
        "If the corrected pincode has a mapped big city, include that city name. "
        "Only Return blank if given pincode doesnot contain numbers"
        "Your response MUST be in strict JSON format with no extra text: "
        "{\"mapped_pincode\": \"correct_pincode\", \"mapped_city\": \"correct_city\"}. "
        "Ensure the pincode is valid and exists. Do NOT return 'unknown' or an empty value."
    )

    #st.write(f"🔹 Sending to Azure: {prompt}")  # ✅ Debugging print

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100
        )

        response_text = response.choices[0].message.content.strip()
        #st.write(f"🔹 Azure Response: {response_text}")  # ✅ Debugging print

        # Clean and parse JSON response
        response_text = response_text.replace("```json", "").replace("```", "").strip()
        parsed_response = json.loads(response_text)

        if isinstance(parsed_response, dict):
            return {
                "mapped_pincode": parsed_response.get("mapped_pincode", "Unknown"),
                "mapped_city": parsed_response.get("mapped_city", "Unknown")
            }
        else:
            return {"mapped_pincode": "Unknown", "mapped_city": "Unknown"}

    except Exception as e:
        #st.write(f"⚠️ Azure API Error: {e}")
        return {"mapped_pincode": "Unknown", "mapped_city": "Unknown"}
    
    
def get_correct_pincode_or_city(pincode):
    """Use Azure LLM to correct pincode and find mapped city."""
    
    prompt = (
        f"The given Indian pincode is '{pincode}', please correct it."
        "If this pincode is too long (more than 6 digits), use only the first 6 digits. "
        "If the pincode is too short (less than 6 digits), try adding a suitable last digit to make it valid. "
        "If the last digit seems incorrect, try changing it to the closest valid number. "
        "Always return the closest correct 6-digit pincode based on known patterns. "
        "If the corrected pincode has a mapped big city, include that city name. "
        "Only Return blank if given pincode doesnot contain numbers"
        "Your response MUST be in strict JSON format with no extra text: "
        "{\"mapped_pincode\": \"correct_pincode\", \"mapped_city\": \"correct_city\"}. "
        "Ensure the pincode is valid and exists. Do NOT return 'unknown' or an empty value."
    )

    #st.write(f"🔹 Sending to Azure: {prompt}")  # ✅ Debugging print

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100
        )

        response_text = response.choices[0].message.content.strip()
        #st.write(f"🔹 Azure Response: {response_text}")  # ✅ Debugging print

        # Clean and parse JSON response
        response_text = response_text.replace("```json", "").replace("```", "").strip()
        parsed_response = json.loads(response_text)

        if isinstance(parsed_response, dict):
            return {
                "mapped_pincode": parsed_response.get("mapped_pincode", "Unknown"),
                "mapped_city": parsed_response.get("mapped_city", "Unknown")
            }
        else:
            return {"mapped_pincode": "Unknown", "mapped_city": "Unknown"}

    except Exception as e:
        #st.write(f"⚠️ Azure API Error: {e}")
        return {"mapped_pincode": "Unknown", "mapped_city": "Unknown"}
    
def get_lat_long(row):
    """Use Azure LLM to get the most probable lat long for a given Indian pincode and city."""
    pincode = row["mapped_pincode"]
    tat_city = row["mapped_city_tat"]
    state_code = row["mapped_state_db"]
    prompt = (
        f'''Given the pincode: '{pincode}', city: '{tat_city}', state code: '{state_code}' in India,
         return the most nearest Latitude and Longitude for the given indian pincode, city, and state code combination.
         prioritize the pincode first, then city and state, as I have seen you gives same lat long for different pincodes
        The response MUST follow strict JSON format with double quotes: 
       {{'Latitude':'latitude,
         'Longitude':'longitude'}}
        You MUST always return a latitude and longitude—never leave it empty or return 'Unknown'.'''
    )

    # return prompt

    response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100
        )

    
    llm_json_output =  response.choices[0].message.content
    try:
        json_match = re.search(r"```json\n([\s\S]+?)\n```", llm_json_output)

        if json_match:
            json_data = json_match.group(1) 
            data = json.loads(json_data) 
            return pd.Series([data['Latitude'], data['Longitude']])

        else:
            return pd.Series([None, None])
        
    except:
        return pd.Series([None , None,])

#### Function to get Lat Long using LocateOne API
def get_lat_lng_locateoneapi(row):
    url = "https://api.getos1.com/locateone/v1/geocode"
    payload = {"data": {
        "address":  f"{row['mapped_pincode']}, {row['mapped_city_db']}, {row['mapped_state_db']}, India",
        "pincode": row['mapped_pincode'],
        "city": row['mapped_city_db'],
        "state": row['mapped_state_db']}}

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "x-api-key": locate_one_api_key
    }

    time.sleep(0.53)  # ~100 calls per min

    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            data = response.json()
            return pd.Series([row['mapped_pincode'],
                              data['result']['geocode']['lat'],
                              data['result']['geocode']['lng']])
        else:
            return pd.Series([row['mapped_pincode'], 'Error', response.text])
    except Exception as e:
        return pd.Series([row['mapped_pincode'], 'Error', str(e)])    

#### Function to get Lat Long using Google Geocoding API
def get_lat_lng_from_geocode_api(row):
    GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"
    input_pincode = row['mapped_pincode']
    address = f"{row['mapped_pincode']}, {row['mapped_city_db']}, {row['mapped_state_db']}, India"
    # address = f"{row['mapped_city_db']}, {row['mapped_state_db']}, India"

    params = {"address": address, "region": "in", "key": google_distance_api_key}

    params = {
    # "components": f"postal_code:{input_pincode}|country:IN",
    "address": address,
    "key": google_distance_api_key
}
    # return params
    try:
        response = requests.get(GEOCODE_URL, params=params, timeout=(3, 5))
        data = response.json()
        return data
        if data.get("status") == "OK" and data["results"]:
            result = data["results"][0]
            formatted_address = result["formatted_address"]
            lat = result["geometry"]["location"]["lat"]
            lng = result["geometry"]["location"]["lng"]
            return pd.Series([input_pincode, formatted_address, lat, lng])
        else:
            return pd.Series([input_pincode, None, None, None])
    except Exception:
        return pd.Series([input_pincode, None, None, None])

# function to create one df for pin, city and state
def unique_pin_city_state(df_input):

    for col in ['pincode', 'destination', 'destination_state', 'pincode_origin', 'origin', 'origin_state']:
        if not col in df_input.columns:
            df_input[col] = ""


    initial_raw_df = df_input.copy()
    unique_od_pair_df = df_input[['pincode', 'destination', 'destination_state', 'pincode_origin', 'origin', 'origin_state']].drop_duplicates().reset_index(drop=True) 

    pincodes_1 = df_input[['pincode', 'destination', 'destination_state']].drop_duplicates()
    pincodes_2 = (df_input[['pincode_origin', 'origin', 'origin_state']].drop_duplicates()
                  .rename(columns={'pincode_origin': 'pincode','origin':'destination','origin_state':'destination_state'}))
    unique_pin_city_state = pd.concat([pincodes_1, pincodes_2], ignore_index=True).drop_duplicates()
    
    return initial_raw_df, unique_od_pair_df, unique_pin_city_state

def parallel_apply_with_progress_osrm(func, data):
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Use tqdm to wrap the executor's map, enabling progress tracking
        result = list(stqdm(executor.map(func, data), total=len(data), desc="Processing"))
    return result 

def parallel_apply_with_progress(func, data):
   
    with ThreadPoolExecutor() as executor:
        # Use tqdm to wrap the executor's map, enabling progress tracking
        result = list(stqdm(executor.map(func, data), total=len(data), desc="Processing"))
    return result 

#  Function to get pincode mapping from internal db and LLM
def pincode_mapping_agent(df_input):
    print('Entered pincode_mapping_agent function')
    df_pincode = get_internal_pincode_db()
    

    rename_map = {
    "pincode_origin": "pincode",
    "origin": "destination",
    "origin_state": "destination_state"}

    for old_col, new_col in rename_map.items():
        if old_col in df_input.columns:
            df_input = df_input.rename(columns={old_col: new_col})


    for col in ["pincode", "destination", "destination_state"]:
        if not col in df_input.columns:
            df_input[col] = ""  
            # column_map["pincode"] = "pincode"

    
    # print("Column Map:", column_map)

    
    
    total_rows = len(df_input)


    df_input['pincode'] = df_input['pincode'].replace('', np.nan)

    df_input['pincode'] = df_input['pincode'].astype('Int64')
    df_input['pincode'] = df_input['pincode'].replace(0, None)
    # return df_input
    df_input["pincode"] = df_input["pincode"].astype(str).str.strip()

    df_input["destination"] = df_input['destination'].astype(str)
    df_input["destination_state"] = df_input['destination_state'].astype(str)

    df_input = df_input.replace('nan','')

    df_input['pincode'] = df_input['pincode'].replace('<NA>', None)
    join_cols = [col for col in ['pincode', 'destination', 'destination_state'] if col in df_input.columns]
    # , 'product_code','product_desc'
    df = df_input[join_cols].drop_duplicates()

    # return df

    df["pincode_flag"] = df["pincode"].swifter.apply(validate_pincode)

    df = df.merge(df_pincode, left_on = "pincode", right_on = "mapped_pincode", how="left")
    # return df
    for index, row in df.iterrows():
        # return row
        import pandas as pd
        if (pd.isna(row["pincode"]) or row["pincode"].strip() == "" or row["pincode"] in ["", "nan", "None"]) and (pd.isna(row["destination"]) or row["destination"].strip() == ""):
            df.at[index, "loc_mapping_type"] = "No Data Provided"
            continue  # Skip further processing for this row


        if not row["pincode_flag"] or pd.isna(row["mapped_city_db"]):  
            # unmapped pincode with internal db
            if pd.isna(row["pincode"]) or row["pincode"] in ["", "nan", "None"] and pd.notna(row['destination']):  
                
                    probable_pincode = get_most_probable_pincode(row) 
                    df.at[index, "mapped_pincode"] = probable_pincode  
                    df.at[index, "loc_mapping_type"] = "Azure OpenAI"

                    # Try mapping the new pincode to internal DB
                    matched_pincode_row = df_pincode[df_pincode["mapped_pincode"] == probable_pincode]

                    if not matched_pincode_row.empty:
                        df.at[index, "mapped_city_db"] = matched_pincode_row["mapped_city_db"].values[0]
                        df.at[index, "mapped_state_db"] = matched_pincode_row["mapped_state_db"].values[0]
                        df.at[index, "mapped_city_tat"] = matched_pincode_row["mapped_city_tat"].values[0]
                        df.at[index, "loc_mapping_type"] = "LLM City-based Pincode Match"
                    else:
                        df.at[index, "loc_mapping_type"] = "LLM Pincode - No Match"

            else:  #pincode present but didn't mapped to internal pincode db
                # **Correct invalid pincode using LLM**
                correction = get_correct_pincode_or_city(row["pincode"])
                df.at[index, "mapped_pincode"] = correction["mapped_pincode"]
                df.at[index, "mapped_city_db"] = correction["mapped_city"]

                if pd.notna(df.at[index, "mapped_city_db"]) and df.at[index, "mapped_city_db"] != "Unknown":
                    city_match = df_pincode[df_pincode["mapped_city_db"].str.lower() == df.at[index, "mapped_city_db"].lower()]
                    
                    if not city_match.empty:
                        df.at[index, "mapped_pincode"] = city_match["mapped_pincode"].values[0]  
                        df.at[index, "mapped_state_db"] = city_match["mapped_state_db"].values[0]  
                        df.at[index, "mapped_city_tat"] = city_match["mapped_city_tat"].values[0]  
                        df.at[index, "loc_mapping_type"] = "City-based Mapping from Internal DB"

                # **Check if the corrected pincode exists in Internal DB**
                matched_corrected_pincode_row = df_pincode[df_pincode["mapped_pincode"] == correction["mapped_pincode"]]

                if not matched_corrected_pincode_row.empty:
                    df.at[index, "mapped_state_db"] = matched_corrected_pincode_row["mapped_state_db"].values[0]
                    df.at[index, "mapped_city_tat"] = matched_corrected_pincode_row["mapped_city_tat"].values[0]
                    df.at[index, "loc_mapping_type"] = "LLM Corrected Pincode Match"
                else:
                    df.at[index, "loc_mapping_type"] = "LLM Pincode - No Match"

        # **Update progress bar**
        progress_percent = int(((index + 1) / total_rows) * 100)
        progress_bar.progress(progress_percent)

        progress_text.text(f"Cleaning cities names... {progress_percent}% completed ({index} rows done) ")
        

    progress_bar.progress(40)


    df.columns = pd.Series(df.columns).where(~pd.Series(df.columns).duplicated(), 
                                        pd.Series(df.columns) + "_" + pd.Series(df.columns).duplicated().cumsum().astype(str))

    # df = df.swifter.apply(fuzzy_match_on_pincode)
    df =  fuzzy_match_on_pincode(df) # Apply fuzzy matching on pincode

    df1 = df[df['Latitude'].isna()]
    df2 = df[df['Latitude'].notna()]
    # Fetch Latitude and Longitude from LLM

    print("Fetching Latitude and Longitude through LLM...")
        
    rows = [row for index, row in df1.iterrows()]

        # Use parallel_apply_with_progress function instead of swifter.apply
    lat_long_output =  parallel_apply_with_progress(lambda row: get_lat_long(row), rows)

    # st.write("llm Response output", lat_long_output)

    latitudes = []
    longitudes = []

    # Iterate through the results array and extract the values
    for res in lat_long_output:
        latitudes.append(res[0])  # Extract latitude
        longitudes.append(res[1]) # Extract longitude


    # # Assign the latitudes and longitudes to the DataFrame columns
    df1['Latitude'] = latitudes
    df1['Longitude'] = longitudes

    # st.write("Added Lat Long to Unmapped", df1)

    df = pd.concat([df1, df2], ignore_index=True )


    if "loc_mapping_type" not in df.columns:
        df["loc_mapping_type"] = ""
    df["loc_mapping_type"] = df["loc_mapping_type"].replace("", pd.NA).fillna("Direct Mapping")
    df = df.fillna("")
    df = df.astype(str).replace("nan", "").replace("None", "").replace("NaN", "")
    df.drop(columns=["pincode_prefix"], errors="ignore", inplace=True)

    # st.write("Processed after lat long:", df.head(10))

    import pandas as pd
    import io
    df_input[['pincode', 'destination', 'destination_state']] = df_input[['pincode', 'destination', 'destination_state']].replace(r'^\s*$', np.nan, regex=True)
    df[['pincode', 'destination', 'destination_state']] = df[['pincode', 'destination', 'destination_state']].replace(r'^\s*$', np.nan, regex=True)

    df['destination_state'] = df['destination_state'].astype(object)
    df['destination'] = df['destination'].astype(object)
    # df = clean_df(df, join_cols)
    # df_input = clean_df(df_input, join_cols)
    df_input['destination_state'] = df_input['destination_state'].astype(object)
    df_input['destination'] = df_input['destination'].astype(object)

    df_input = df_input.fillna('0')
    df  = df.fillna('0')
    df_input_processed = df_input.merge(df, how='left', on = join_cols)

    df_input_processed = df_input_processed.fillna('0')

    return df_input_processed


# Separate pincodwe mapping at origin & destination level 
def origin_destination_mapping(df_input_processed, raw_od_pair_df):

    raw_od_pair_df = raw_od_pair_df.fillna(0)
    raw_od_pair_df = raw_od_pair_df.replace('','0')

    # st.write('start of origin_destination_mapping raw_od_pair_df',raw_od_pair_df.dtypes, raw_od_pair_df)
    
    for col in ['pincode','pincode_origin']:

        raw_od_pair_df[col] = raw_od_pair_df[col].astype('Int64')
        raw_od_pair_df[col] = raw_od_pair_df[col].astype(object)

    for col in ['pincode', 'destination', 'destination_state', 'pincode_origin', 'origin', 'origin_state']:
        raw_od_pair_df[col] = raw_od_pair_df[col].apply(lambda x: int(x) if isinstance(x, float) and x.is_integer() else x).astype(str).str.strip()

    for col in ['pincode', 'destination', 'destination_state']:
        df_input_processed[col] = df_input_processed[col].astype(str).str.strip()

    # st.write('in origin_destination_mapping raw_od_pair_df',raw_od_pair_df)
    # st.write('in origin_destination_mapping df_input_processed',raw_od_pair_df.dtypes, df_input_processed)

    raw_od_pair_df = raw_od_pair_df.merge(df_input_processed, how='left', 
                            left_on=['pincode', 'destination', 'destination_state'], right_on = ['pincode', 'destination', 'destination_state'])

    # st.write('in origin_destination_mapping raw_od_pair_df',raw_od_pair_df)
    raw_od_pair_df = raw_od_pair_df.merge(df_input_processed, how='left', 
                            left_on=['pincode_origin', 'origin', 'origin_state'], right_on = ['pincode', 'destination', 'destination_state'], suffixes=('', '_origin_1'))
    
    # st.write('raw_od_pair_df columns', raw_od_pair_df.columns)
    # raw_od_pair_df = raw_od_pair_df.drop(columns=['pincode_origin_1', 'origin_1', 'origin_state_1'])
    # raw_od_pair_df = raw_od_pair_df.drop(columns=['pincode_origin_1', 'city_origin_1', 'state_origin_1'])
    
    return raw_od_pair_df

# Creating a custom parallel process function for OSRM API
def parallel_apply_with_progress_osrm(func, data):
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Use tqdm to wrap the executor's map, enabling progress tracking
        result = list(stqdm(executor.map(func, data), total=len(data), desc="Processing"))
    return result 

def parallel_apply_with_progress(func, data):
   
    with ThreadPoolExecutor() as executor:
        # Use tqdm to wrap the executor's map, enabling progress tracking
        result = list(stqdm(executor.map(func, data), total=len(data), desc="Processing"))
    return result 


# Fetching distance from OSRM API using OD Lat and Long 
def get_distance_from_osrm(row, mode='driving'):
    try:
        origin_lng = row['Longitude_origin_1']
        origin_lat = row['Latitude_origin_1']

        destination_lng = row['Longitude']
        destination_lat = row['Latitude']

        # print(abs(float(origin_lng) -  float(destination_lng)))
        # print(abs(float(origin_lat) -  float(origin_lat)))

        if abs(float(origin_lng) -  float(destination_lng)) < 0.05 and abs(float(origin_lat) -  float(origin_lat)) <0.05:
            return pd.Series([20])
        
        else:

            # """
            # Function to get distance from OSRM API.
            # """
            base_url = 'http://router.project-osrm.org/route/v1'

            profile = mode  # can be 'driving', 'walking', 'cycling'
            coordinates = f"{origin_lng},{origin_lat};{destination_lng},{destination_lat}"

            url = f"{base_url}/{profile}/{coordinates}?overview=false"

            # print(url)
            response = requests.get(url)
            # return response
            try: 
                data = response.json()  # This line could raise a JSONDecodeError
            # return data
            except json.JSONDecodeError:
                return pd.Series([distance_in_km])

                # Check if the response contains valid data
            if 'routes' in data and len(data['routes']) > 0:
                    # Distance is in meters
                    distance_in_meters = data['routes'][0]['distance']

                    # Convert meters to kilometers
                    distance_in_km = distance_in_meters / 1000
                    return pd.Series([distance_in_km])

            else:
                return pd.Series([None])
            
    except Exception as e:
        return pd.Series([None])
    

#### Function to get google distances using Lat Long from LocateOne API
def get_distance_from_google(row, mode='driving'):

    # Create origin and destination distance using pincode, city state
    # origin = f"{row['mapped_pincode_origin_1']}, {row['mapped_city_db_origin_1']}, India"
    # destination = f"{row['mapped_pincode']}, {row['mapped_city_db']}, India"

    # Create origin and destination distance using Lat Long
    origin = f"{row['Latitude_origin_1']}, {row['Longitude_origin_1']}"
    destination = f"{row['Latitude']}, {row['Longitude']}"

    base_url = "https://maps.googleapis.com/maps/api/distancematrix/json"

    params = {
        "origins": origin,
        "destinations": destination,
        "mode": mode,           # driving, walking, bicycling, transit
        "units": "metric",
        "region": "in",         # bias towards India
        "key": google_distance_api_key
    }

    try:
        response = requests.get(base_url, params=params, timeout=10)
        data = response.json()
    except Exception as e:
        return pd.Series([None, None, None])

    if data.get("status") == "OK":
        element = data["rows"][0]["elements"][0]
        if element.get("status") == "OK":
            origin_address = data.get("origin_addresses", [None])[0]
            destination_address = data.get("destination_addresses", [None])[0]
            distance_km = element["distance"]["value"] / 1000  # meters to km
            return pd.Series([origin_address, destination_address, distance_km])
        else:
            return pd.Series([None, None, None])
    else:
        return pd.Series([None, None, None])


import requests
import pandas as pd
from math import ceil


def get_distance_batched(df,api_key, mode="driving", batch_size=10):

    base_url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    results = []

    total_rows = len(df)
    num_batches = ceil(len(df) / batch_size)

    #### Streamlit UI elements
    progress_bar = st.progress(0)
    progress_text = st.empty()

    rows_processed = 0

    for b in range(num_batches):

        chunk = df.iloc[b*batch_size : (b+1)*batch_size]

        # Build coordinate lists
        origins = [f"{row['Latitude_origin_1']},{row['Longitude_origin_1']}" 
                   for _, row in chunk.iterrows()]
        destinations = [f"{row['Latitude']},{row['Longitude']}" 
                        for _, row in chunk.iterrows()]

        # Keep pincode identifiers aligned with same indexes
        origin_pins = chunk["mapped_pincode_origin_1"].tolist()
        destination_pins = chunk["mapped_pincode"].tolist()

        params = {
            "origins": "|".join(origins),
            "destinations": "|".join(destinations),
            "mode": mode,
            "units": "metric",
            "region": "in",
            "key": api_key        }


        try:
            response = requests.get(base_url, params=params, timeout=10)
            data = response.json()
            
        except:
            # Return None for all rows in this chunk
            for i in range(len(chunk)):
                results.append([
                    origin_pins[i],
                    destination_pins[i],
                    None, None, None
                ])

            rows_processed += len(chunk)
            progress_bar.progress(rows_processed / total_rows)
            progress_text.text(f"Fetched distance for {rows_processed} rows  out of {total_rows} rows")
            continue

        # If Google API fails for the whole request
        if data.get("status") != "OK":
            for i in range(len(chunk)):
                results.append([
                    origin_pins[i],
                    destination_pins[i],
                    None, None, None
                ])
            rows_processed += len(chunk)
            progress_bar.progress(rows_processed / total_rows)
            progress_text.text(f"Fetched distance for {rows_processed} rows  out of {total_rows} rows")
            continue

        rows = data.get("rows", [])
        origin_addr_list = data.get("origin_addresses", [])
        dest_addr_list = data.get("destination_addresses", [])

        # Extract diagonal elements (row i → destination i)
        for i in range(len(chunk)):

            try:
                element = rows[i]["elements"][i]
            except:
                results.append([origin_pins[i], destination_pins[i], None, None, None])
                continue

            if element.get("status") == "OK":
                distance_km = element["distance"]["value"] / 1000

                results.append([
                    origin_pins[i],
                    destination_pins[i],
                    origin_addr_list[i],
                    dest_addr_list[i],
                    distance_km
                ])
            else:
                results.append([
                    origin_pins[i],
                    destination_pins[i],
                    None, None, None
                ])

        ### Row counter update
        rows_processed += len(chunk)
        progress_bar.progress(rows_processed / total_rows)
        progress_text.text(f"Fetched distance for {rows_processed} rows out of {total_rows} rows")

    progress_text.text("Completed.")

    # Final output dataframe
    return pd.DataFrame(results, columns=[
        'mapped_pincode_origin_1','mapped_pincode','origin_address','destination_addresses', 'google_distance_km'])


# Joining OD distace data with initial raw df
def join_raw_df(raw_df, initial_data_OD_lat_long):
    
    raw_df = raw_df.fillna(0)
    raw_df = raw_df.replace('','0')

    for col in ['pincode','pincode_origin']:

        raw_df[col] = raw_df[col].astype('Int64')
        raw_df[col] = raw_df[col].astype(object)

    # raw_df['pincode_origin'] = raw_df['pincode_origin'].astype('Int64')
    # raw_df['pincode_origin'] = raw_df['pincode_origin'].astype(object)

    for col in ['pincode', 'destination', 'destination_state', 'pincode_origin', 'origin', 'origin_state']:
        raw_df[col] = raw_df[col].apply(lambda x: int(x) if isinstance(x, float) and x.is_integer() else x).astype(str).str.strip()
        initial_data_OD_lat_long[col] = initial_data_OD_lat_long[col].apply(lambda x: int(x) if isinstance(x, float) and x.is_integer() else x).astype(str).str.strip()

    final_df = raw_df.merge(initial_data_OD_lat_long, how='left', on = ['pincode', 'destination', 'destination_state', 'pincode_origin', 'origin', 'origin_state'])
    return final_df

###### Pincode functions - Ends ######


##### Py

uploaded_file = st.file_uploader("Upload RFQ Sample File", type=["xlsx", "xls","csv"])

if 'column_names' not in st.session_state:
    st.session_state.column_names = []
if 'column_map' not in st.session_state:
    st.session_state.column_map = {}


stqdm.pandas()
if uploaded_file:
    file_name = uploaded_file.name.lower()
    if file_name.endswith(".csv"):
        # Preview without header
        df_preview = pd.read_csv(uploaded_file, header=None)
        sheet_names =  os.path.splitext(uploaded_file.name)[0]
        # st.write('sheet_names',sheet_names)
        sheet_name = st.selectbox("Select Sheet", sheet_names)
        header_row = st.number_input("Select Header Row (0-indexed)", min_value=0, max_value=len(df_preview)-1, value=0)
        # Re-read with header
        uploaded_file.seek(0)  # reset pointer
        df_input = pd.read_csv(uploaded_file, header=header_row)

    elif file_name.endswith((".xlsx", ".xls")):
        # Read Excel and get sheet names
        xls = pd.ExcelFile(uploaded_file)
        sheet_name = st.selectbox("Select Sheet", xls.sheet_names)
        # Preview without header
        df_preview = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        header_row = st.number_input("Select Header Row (0-indexed)", min_value=0, max_value=len(df_preview)-1, value=0)
        # Read with proper header
        df_input = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)

    st.write("Data with Selected Header:", df_input.head(10))
    

    st.session_state.column_names = list(df_input.columns)

    column_names = st.session_state.column_names
    # st.write("Identified columns:", st.session_state.column_names)

    st.markdown("""
        <style>
        div.stButton > button:first-child {
            background-color: #007BFF;  /* Bootstrap Primary Blue */
            color: white;
            padding: 0.75em 1.5em;
            font-size: 1.2em;
            font-weight: bold;
            border-radius: 0.5em;
            border: none;
            transition: background-color 0.3s;
        }
        div.stButton > button:first-child:hover {
            background-color: #0056b3;
            cursor: pointer;}
        </style>""", unsafe_allow_html=True)

    if st.button("Auto-Map Columns Using AI"):
        st.success("Proceeding to map columns...")
        st.session_state.llm_response = get_llm_suggestions(column_names)
        st.session_state.column_map = {} 
        # st.rerun()
    
    if "llm_response" in st.session_state:
        llm_response = st.session_state.llm_response
        expected_columns = ["origin", "origin_state","destination", "destination_state","vehicle_type","load_mt"]

        for col in expected_columns:
            suggested_col = llm_response.get(col, "Not found")
            st.session_state.column_map[col] = st.selectbox(f"Select column for {col}", [None] + column_names, 
            index=(column_names.index(suggested_col) + 1) if suggested_col in column_names else 0, key=f"alt_{col}")

    st.write("## Column Mapping Summary")
    column_map = st.session_state.column_map
    st.write(column_map) 
    clean_mapping = {v: k for k, v in column_map.items() if v is not None}
    df_input = df_input.rename(columns=clean_mapping)
    raw_first_df = df_input.copy()
    st.write("mapped columns preview", df_input.head(10))

    # city_cluster_mapping, ODVTFTL_pricing, OCDVTFTL_pricing, RFQ_FTL_Pricing
    if st.button("Refresh Pricing Data"):
        city_cluster_mapping, ODVT_FTL_pricing, OCDVT_FTL_pricing, RFQ_FTL_Pricing = load_data(refresh_time=datetime.now())
        st.write('### Pricing Data Refreshed')
        st.write('Unique ODVT pairs in FTL pricing data : ', ODVT_FTL_pricing.shape[0])
        st.write('Unique OCDVT pairs in FTL pricing data : ', OCDVT_FTL_pricing.shape[0])
        st.write('Unique OCDCVT pairs in RFQ or FTL pricing data : ', RFQ_FTL_Pricing.shape[0])
        st.write('Unique Cities in City-Cluster mapping data : ', city_cluster_mapping.shape[0])
    else:
        city_cluster_mapping, ODVT_FTL_pricing, OCDVT_FTL_pricing, RFQ_FTL_Pricing  = load_data()

    print('odvt FTL Pricing dataframe shape', RFQ_FTL_Pricing.shape)
    print('ocdvt FTL Pricing dataframe shape', RFQ_FTL_Pricing.shape)
    print('ocdcvt FTL and RFQ Pricing dataframe shape', RFQ_FTL_Pricing.shape)
    # print('mapping_type count',RFQ_FTL_Pricing.groupby('mapping_type').size())
    print('City-Cluster mapping data size', city_cluster_mapping.shape)

    if st.button("Fetch Pricing for Live RFQ's"):
        progress_bar = st.progress(0)
        progress_text = st.empty()
        with st.spinner("Processing..."):

            ### function flow to get distance betwwen given sets of pincodes
          
            
            initial_raw_df, unique_od_pair_df, unique_pin_city_state_df = unique_pin_city_state(df_input)
            # st.write('unique_pin_city_state_df',unique_pin_city_state_df.head(50))
            df_input_processed = pincode_mapping_agent(unique_pin_city_state_df)

            # st.write('pincode mapped df', df_input_processed.head(50))
            # st.write('unique OD mapped df', unique_od_pair_df.head(50))

            initial_data_OD_lat_long = origin_destination_mapping(df_input_processed, unique_od_pair_df)

            # st.write('OD mapped df', initial_data_OD_lat_long.head(50))

            # st.write("Fetching distance between origin and destination through OSRM API. Please wait...")
            # rows = [row for index, row in initial_data_OD_lat_long.iterrows()]
            # distance_km = parallel_apply_with_progress_osrm(lambda row: get_distance_from_osrm(row), rows)
            # # If there are nulls in distance_km then check for nulls in lat and long and then try to change parallel_apply_with_progress_osrm function 
            # distance_df = pd.DataFrame([item.tolist() if item is not None else [None] * 1
            #             for item in distance_km], columns=['distance_km'])
            # initial_data_OD_lat_long['distance_km'] = distance_df

            ###### Calculating distance using Google API
            st.write("Fetching distance between origin and destination through Google API. Please wait...")
            
            # rows = [row for index, row in initial_data_OD_lat_long.iterrows()]
            # google_distance_km = parallel_apply_with_progress_osrm(lambda row: get_distance_from_google(row), rows)
            # google_distance_df = pd.DataFrame([item.tolist() if item is not None else [None] * 1
            #             for item in google_distance_km], columns=['origin_address','destination_addresses', 'google_distance_km'])
            # initial_data_OD_lat_long[['origin_address','destination_addresses', 'google_distance_km']] = google_distance_df

            # st.write('Data size pre google distance', initial_data_OD_lat_long.shape)

            google_distance_df = get_distance_batched(df=initial_data_OD_lat_long, api_key=google_distance_api_key, batch_size=10)
            
            
            initial_data_OD_lat_long = initial_data_OD_lat_long.merge(google_distance_df, on = ['mapped_pincode_origin_1','mapped_pincode'],how = 'left')

            # st.write('Data size post google distance', initial_data_OD_lat_long.shape)

            final_df = join_raw_df(initial_raw_df, initial_data_OD_lat_long)
            # st.write('initial_data_OD_lat_long df', initial_data_OD_lat_long)
            final_df = final_df.fillna('0')
            final_df = final_df.replace('<NA>','0')
            final_df = final_df.replace('nan','0')

            # st.write('distance mapped df', final_df)

            print('Pincode Agent ends')


            final_df = (final_df.rename(columns = {'origin':'input_origin','origin_state':'input_origin_state','destination':'input_destination',
                                           'destination_state':'input_destination_state',
                                           'vehicle_type':'input_vehicle_type',	'load_mt':'input_load_mt'}))
            
            state_cols = ['origin_state', 'destination_state']

            if any(col in raw_first_df.columns for col in state_cols): # State column present in raw data
                drop_cols= ['pincode','pincode_origin','pincode_flag','mapped_city_tat',
                    'pincode_origin_1','destination_origin_1',	'destination_state_origin_1', 'pincode_flag_origin_1','mapped_city_tat_origin_1',
                    'loc_mapping_type','loc_mapping_type_origin_1']
            
                final_df = (final_df.drop(columns = drop_cols))

            else: # State column not present in raw data
                drop_cols= ['pincode','pincode_origin','input_origin_state','input_destination_state','pincode_flag','mapped_city_tat',
                    'pincode_origin_1','destination_origin_1',	'destination_state_origin_1', 'pincode_flag_origin_1','mapped_city_tat_origin_1',
                    'loc_mapping_type','loc_mapping_type_origin_1']
            
                final_df = (final_df.drop(columns = drop_cols))

            final_df['origin'] = np.where(final_df['mapped_city_db_origin_1'].isna() | (final_df['mapped_city_db_origin_1'] == ''),final_df['input_origin'], final_df['mapped_city_db_origin_1'])
            final_df['destination'] = np.where(final_df['mapped_city_db'].isna() | (final_df['mapped_city_db'] == ''),final_df['input_destination'], final_df['mapped_city_db'])
            final_df['vehicle_type'] = final_df['input_vehicle_type']
            final_df['load_mt'] = final_df['input_load_mt']

            df_input = final_df
             ## Assign pincode agent output for pricing data input
            # st.write('pincode cleaned output', df_input)

            #### Lane pricing mapping starts here ---- 
            df_input['origin'] = df_input['origin'].str.lower()
            df_input['destination'] = df_input['destination'].str.lower()
            df_input['vehicle_type'] = df_input['vehicle_type'].str.lower()
            df_input = df_input.drop_duplicates()

            state_cols = ['origin_state', 'destination_state']
            # if any(col in raw_first_df.columns for col in state_cols):
            state_cols_input = ['input_origin_state','input_destination_state']
            for col in state_cols_input:
                if col not in df_input.columns:
                    df_input[col] = ''

        # Mapping clusters to cities in rfq data based on cities & states
            df_input['origin_state'] = np.where(df_input['input_origin_state'].isna() | (df_input['input_origin_state'].isin([None, '', 0,'0'])),df_input['mapped_state_db_origin_1'],
            df_input['input_origin_state'])
            df_input['destination_state'] = np.where(df_input['input_destination_state'].isna() | (df_input['input_destination_state'].isin([None, '', 0,'0'])),df_input['mapped_state_db'],
            df_input['input_destination_state'])

            df_input['origin_state'] = df_input['origin_state'].str.upper()
            df_input['destination_state'] = df_input['destination_state'].str.upper()

                # city_cluster_mapping_unique = (city_cluster_mapping.sort_values(['city','state'])
                #                                .drop_duplicates(subset=['city','state'], keep='first'))

                # df_input_cluster = df_input.merge(city_cluster_mapping_unique, left_on=['origin','origin_state'], right_on=['city','state'], how = 'left', suffixes=('', '_origin'))
                # df_input_cluster = df_input_cluster.merge(city_cluster_mapping_unique, left_on=['destination','destination_state'], right_on=['city','state'], how = 'left', suffixes=('', '_destination'))
                

            # Mapping clusters to citites in rfq data based on cities
            # else: 
            #     df_input['origin_state'] = df_input['mapped_state_db_origin_1']
            #     df_input['destination_state'] =  df_input['mapped_state_db']
                # city_cluster_mapping_unique = (city_cluster_mapping.sort_values('city').drop_duplicates(subset=['city'], keep='first'))
                # df_input_cluster = df_input.merge(city_cluster_mapping_unique, left_on=['origin'], right_on=['city'], how = 'left', suffixes=('', '_origin'))
                # df_input_cluster = df_input_cluster.merge(city_cluster_mapping_unique, left_on=['destination'], right_on=['city'], how = 'left', suffixes=('', '_destination'))
                
             
            city_cluster_mapping_unique = (city_cluster_mapping.sort_values(['city','state'])
                                               .drop_duplicates(subset=['city','state'], keep='first'))
            
            
            ## city column names: input_origin, input_destination, origin, destination

            # ---------- PRIMARY ORIGIN MAPPING ---------- "input_origin", "origin_state"
            primary = df_input.reset_index().merge(
                city_cluster_mapping_unique,
                left_on=["input_origin", "origin_state"],
                right_on=["city", "state"],
                how="left"
            ).set_index("index")[["cluster", "city_code"]]

            fallback = df_input.reset_index().merge(
                city_cluster_mapping_unique,
                left_on=["origin", "origin_state"],
                right_on=["city", "state"],
                how="left"
            ).set_index("index")[["cluster", "city_code"]]

            df_input["city_code_origin"] = primary["city_code"].combine_first(fallback["city_code"])
            df_input["cluster_origin"] = primary["cluster"].combine_first(fallback["cluster"])
            

            # df.loc[df["cluster_origin"].isna(), "cluster_origin"] = fallback["cluster"]
            # df.loc[df["cluster_origin"].isna(), "city_code_origin"] = fallback["city_code"]


            # ---------- PRIMARY Destination MAPPING ---------- "input_destination", "destination_state"
            primary_d = df_input.reset_index().merge(
                city_cluster_mapping_unique,
                left_on=["input_destination", "destination_state"],
                right_on=["city", "state"],
                how="left"
            ).set_index("index")[["cluster", "city_code"]]

            fallback_d = df_input.reset_index().merge(
                city_cluster_mapping_unique,
                left_on=["destination", "destination_state"],
                right_on=["city", "state"],
                how="left"
            ).set_index("index")[["cluster", "city_code"]]

            df_input["city_code_destination"] = primary_d["city_code"].combine_first(fallback_d["city_code"])
            df_input["cluster_destination"] = primary_d["cluster"].combine_first(fallback_d["cluster"])
            

            df_input_cluster = df_input.copy()

            # st.write('df_input_cluster after primary and fallback mapping', df_input_cluster)

            

            # df_input_cluster = df_input.merge(city_cluster_mapping_unique, left_on=['origin','origin_state'], right_on=['city','state'], how = 'left', suffixes=('', '_origin'))
            # df_input_cluster = df_input_cluster.merge(city_cluster_mapping_unique, left_on=['destination','destination_state'], right_on=['city','state'], how = 'left', suffixes=('', '_destination'))
                

            #### create cluster mapping unique at city level   
            # city_cluster_mapping_unique = (city_cluster_mapping.sort_values('city').drop_duplicates(subset=['city'], keep='first'))

            # df_input_cluster = df_input_cluster.rename(columns = {'cluster':'cluster_origin','city_code':'city_code_origin'})

            df_input_cluster['load_mt'] = df_input_cluster['load_mt'].astype(float)
            # st.write('df_input_cluster raw',df_input_cluster)
            # st.write("rfq data cluster joined", df_input_cluster.head(10))
            
            ######## Add Major city function here, for missing clusters
            # if any(col in raw_first_df.columns for col in state_cols):
            missing_cluster_oc_cities = df_input_cluster[df_input_cluster['cluster_origin'].isna()][['origin','origin_state']].rename(columns = {'origin':'city','origin_state':'state'})
            missing_cluster_dc_cities = df_input_cluster[df_input_cluster['cluster_destination'].isna()][['destination','destination_state']].rename(columns = {'destination':'city','destination_state':'state'})

            # else: 
            #     missing_cluster_oc_cities = df_input_cluster[df_input_cluster['cluster_origin'].isna()][['origin']].rename(columns = {'origin':'city'})
            #     missing_cluster_dc_cities = df_input_cluster[df_input_cluster['cluster_destination'].isna()][['destination']].rename(columns = {'destination':'city'})

            all_missing_cities = pd.concat([missing_cluster_oc_cities, missing_cluster_dc_cities])
            all_missing_cities = all_missing_cities.drop_duplicates().reset_index(drop = True)

            # st.write('missing cities df',all_missing_cities)
            
            rows = [row for index, row in all_missing_cities.iterrows()]
            mapped_major_city = parallel_apply_with_progress(lambda row: get_nearest_major_city(row), rows)

            mapped_major_city_df = pd.DataFrame([item.tolist() if item is not None else [None] * 3
                    for item in mapped_major_city], columns=['missing_cluster_city','missing_state','mapped_major_city'])
            

            # st.write('mapped_major_city_df',mapped_major_city_df)
            ### fetch clusters for LLM mapped cities
            mapped_major_city_df['mapped_major_city'] = mapped_major_city_df['mapped_major_city'].str.lower()
            major_city_cluster_mapped = mapped_major_city_df.merge(city_cluster_mapping_unique, left_on=['mapped_major_city','missing_state'], right_on=['city','state'], how = 'left', suffixes=('', '_appended'))
            major_city_cluster_mapped = major_city_cluster_mapped.rename(columns = {'cluster':'cluster_appended','city_code':'city_code_appended'})
            major_city_cluster_mapped = major_city_cluster_mapped[['missing_cluster_city','missing_state','mapped_major_city','cluster_appended','city_code_appended']]
            
            # st.write('major_city_cluster_mapped',major_city_cluster_mapped)

            ### join with input data
            df_input_cluster = df_input_cluster.merge(major_city_cluster_mapped, left_on=['origin','origin_state'], right_on=['missing_cluster_city','missing_state'], how = 'left', suffixes=('', '_origin'))
            df_input_cluster = df_input_cluster.merge(major_city_cluster_mapped, left_on=['destination','destination_state'], right_on=['missing_cluster_city','missing_state'], how = 'left', suffixes=('', '_destination'))
            
            # st.write('df_input_cluster cluster added', df_input_cluster)

            ### replace missing clusters with LLM mapped cities clusters
            # df_input_cluster['city'] = df_input_cluster['city'].fillna(df_input_cluster['mapped_major_city'])
            df_input_cluster['city_code_origin'] = df_input_cluster['city_code_origin'].fillna(df_input_cluster['city_code_appended'])
            df_input_cluster['cluster_origin'] = df_input_cluster['cluster_origin'].fillna(df_input_cluster['cluster_appended'])

            # df_input_cluster['city_destination'] = df_input_cluster['city_destination'].fillna(df_input_cluster['mapped_major_city_destination'])
            df_input_cluster['city_code_destination'] = df_input_cluster['city_code_destination'].fillna(df_input_cluster['city_code_appended_destination'])
            df_input_cluster['cluster_destination'] = df_input_cluster['cluster_destination'].fillna(df_input_cluster['cluster_appended_destination'])

            # st.write('## df_input_cluster cluster added',df_input_cluster)

            # Mapping odvtmt pricing on rfq data 
            # odvtmt_join_cols = ["origin","destination","vehicle_type","load_mt"]
            df_input_cluster_cols = ["origin","destination","cluster_origin","cluster_destination","vehicle_type","load_mt"]
            # ODVT_FTL_pricing, OCDVT_FTL_pricing, RFQ_FTL_Pricing

            og_columns = df_input_cluster.columns

            # st.write('df_input_cluster.columns', df_input_cluster.columns)
            # t.write('ODVT_FTL_pricing.columns', ODVT_FTL_pricing.columns)
            ###### Joining input df with ODVT FTL Pricing Data 

            # if any(col in raw_first_df.columns for col in state_cols):
            ODVT_FTL_pricing['origin'] = ODVT_FTL_pricing['origin'].str.lower()
            ODVT_FTL_pricing['destination'] = ODVT_FTL_pricing['destination'].str.lower()
            ODVT_FTL_pricing['SVT'] = ODVT_FTL_pricing['SVT'].str.lower()

            OCDVT_FTL_pricing['cluster_origin'] = OCDVT_FTL_pricing['cluster_origin'].str.lower()
            OCDVT_FTL_pricing['destination'] = OCDVT_FTL_pricing['destination'].str.lower()
            OCDVT_FTL_pricing['SVT'] = OCDVT_FTL_pricing['SVT'].str.lower()

            RFQ_FTL_Pricing['final_origin_cluster'] = RFQ_FTL_Pricing['final_origin_cluster'].str.lower()
            RFQ_FTL_Pricing['final_destination_cluster'] = RFQ_FTL_Pricing['final_destination_cluster'].str.lower()
            RFQ_FTL_Pricing['final_VT'] = RFQ_FTL_Pricing['final_VT'].str.lower()

            df_input_cluster_joined1 = df_input_cluster.merge(ODVT_FTL_pricing, left_on=["origin","origin_state","destination","destination_state","vehicle_type"], 
                                right_on=["origin","origin_state_code","destination","destination_state_code","SVT"], how = 'left')
            ###### Joining input df with OCDVT FTL Pricing Data 
            # st.write('df_input_cluster_joined1', df_input_cluster_joined1)
            # st.write('OCDVT_FTL_pricing', OCDVT_FTL_pricing)

            df_input_cluster_joined2 = df_input_cluster_joined1.merge(OCDVT_FTL_pricing, left_on=["cluster_origin","destination","destination_state","vehicle_type"], 
                                    right_on=["cluster_origin","destination","destination_state_code","SVT"], how = 'left')

            # else:
            #     ### Create unique df at ODVT level
            #     ODVT_FTL_pricing = (ODVT_FTL_pricing.sort_values(['origin','destination'])
            #                                    .drop_duplicates(subset=["origin","destination","SVT"], keep='first'))
            #     df_input_cluster_joined1 = df_input_cluster.merge(ODVT_FTL_pricing, left_on=["origin","destination","vehicle_type"], 
            #                         right_on=["origin","destination","SVT"], how = 'left')
                
            #     ###### Joining input df with OCDVT FTL Pricing Data 

            #     ### Create unique df at OCDVT level
            #     OCDVT_FTL_pricing = (OCDVT_FTL_pricing.sort_values(['cluster_origin','destination'])
            #                                    .drop_duplicates(subset=["cluster_origin","destination","SVT"], keep='first'))
            #     df_input_cluster_joined2 = df_input_cluster_joined1.merge(OCDVT_FTL_pricing, left_on=["cluster_origin","destination","vehicle_type"], 
            #                         right_on=["cluster_origin","destination","SVT"], how = 'left')
                
            ###### Joining input df with OCDCVT FTL & RFQ Pricing Data 

            rfq_samples_matched = df_input_cluster_joined2.merge(RFQ_FTL_Pricing, left_on=["cluster_origin","cluster_destination","vehicle_type"], 
                                 right_on=["final_origin_cluster",	"final_destination_cluster",	"final_VT"], how = 'left')

            # st.write('rfq_samples_matched df final',rfq_samples_matched)

            # odvtmt_level_joined_city = rfq_samples_city_matched_ODVTMT[rfq_samples_city_matched_ODVTMT['final_origin'].notnull()]
            # odvt_level_joined_city = rfq_samples_city_matched_ODVT[rfq_samples_city_matched_ODVT['final_origin'].notnull()]


            # Cluster level match for remaining RFQs
            # unmatched_rfq_on_city_ODVTMT = rfq_samples_city_matched_ODVTMT[rfq_samples_city_matched_ODVTMT['final_origin'].isna()][og_columns]
            # unmatched_rfq_on_city_ODVT = rfq_samples_city_matched_ODVT[rfq_samples_city_matched_ODVT['final_origin'].isna()][og_columns]


            # # Joining on ODVTMT level - cluster level
            # rfq_samples_cluster_matched_ODVTMT = unmatched_rfq_on_city_ODVTMT.merge(RFQ_FTL_Pricing[RFQ_FTL_Pricing['mapping_type']=='cluster'], left_on=["cluster_origin", "cluster_destination","vehicle_type","load_mt"],
            #                             right_on=['final_origin', 'final_destination', 'final_VT', 'final_tonnage'], how = 'left')

            # # Joining on ODVT level - cluster level
            # rfq_samples_cluster_matched_ODVT = unmatched_rfq_on_city_ODVT.merge(RFQ_FTL_Pricing[RFQ_FTL_Pricing['mapping_type']=='cluster'], left_on=["cluster_origin", "cluster_destination","vehicle_type"],
            #                                 right_on=['final_origin', 'final_destination', 'final_VT'], how = 'left')

            
            summary = {
                'total_rfqs': df_input_cluster.shape[0],
                'OCDVT Level Matched': rfq_samples_matched[rfq_samples_matched['ODVT_trips'].notnull()][df_input_cluster_cols].drop_duplicates().shape[0],
                'OCDVT Level Matched': rfq_samples_matched[rfq_samples_matched['OCDVT_trips'].notnull()][df_input_cluster_cols].drop_duplicates().shape[0],
                'OCDCVT Level matched': rfq_samples_matched[rfq_samples_matched['final_origin_cluster'].notnull()][df_input_cluster_cols].drop_duplicates().shape[0]
            }

            # st.write('rfq_samples_matched', rfq_samples_matched, summary)
            # rfq_odvtmt_output_df = pd.concat([odvtmt_level_joined_city,rfq_samples_cluster_matched_ODVTMT], ignore_index=True)
            rfq_odvt_output_df = rfq_samples_matched
            # pd.concat([odvt_level_joined_city,rfq_samples_cluster_matched_ODVT], ignore_index=True)
    
            # rfq_odvtmt_output_df = rfq_odvtmt_output_df.drop(columns=['city','city_destination'])
            # st.write('rfq_odvt_output_df.columns', rfq_odvt_output_df.columns)
            # st.write('rfq_odvt_output_df sample data', rfq_odvt_output_df.head(10))
            # print(rfq_odvt_output_df.columns)

            # if any(col in raw_first_df.columns for col in state_cols):
            rfq_odvt_output_df = rfq_odvt_output_df.drop(columns=[
                # 'city','state', 'cluster_new','city_destination','state_destination', 'cluster_new_destination',
                    'SVT_x','SVT_y','cluster_origin_x', 'cluster_destination_x', 
                    'origin_state_code',	'destination_state_code_x', 'destination_state_code_y',
                    'SVT','final_origin_cluster','final_destination_cluster','final_VT',
                    'missing_cluster_city',	'missing_state','cluster_appended',	'city_code_appended',
                    'missing_cluster_city_destination','missing_state_destination','cluster_appended_destination', 'city_code_appended_destination'])
            
            if not any(col in raw_first_df.columns for col in state_cols): # if not state column was given as input then drop null columns
                rfq_odvt_output_df = rfq_odvt_output_df.drop(columns=['input_origin_state',	'input_destination_state'])

            elif 'origin_state' not in raw_first_df.columns:
                rfq_odvt_output_df = rfq_odvt_output_df.drop(columns=['input_origin_state'])

            elif 'destination_state' not in raw_first_df.columns:
                rfq_odvt_output_df = rfq_odvt_output_df.drop(columns=['input_destination_state'])



            # else:
            #     rfq_odvt_output_df = rfq_odvt_output_df.drop(columns=['city','cluster_new','city_destination',
            #             'cluster_new_destination','SVT_x','SVT_y','cluster_origin_x', 'cluster_destination_x', 
            #             'SVT','final_origin_cluster','final_destination_cluster','final_VT',
            #             'missing_cluster_city',	'cluster_appended',	'city_code_appended',
            #             'missing_cluster_city_destination','cluster_appended_destination',	'city_code_appended_destination'])

            rfq_odvt_output_df = rfq_odvt_output_df.rename(columns={'cluster_origin_y':'rfq_cluster_origin',
                    'cluster_destination_x':'rfq_cluster_destination','rfq_vt':'rfq_vehicle_type'})
            
            rfq_odvt_output_df['vehicle_type'] = rfq_odvt_output_df['vehicle_type'].astype(str).str.upper()
            rfq_odvt_output_df['rfq_vehicle_type'] = rfq_odvt_output_df['rfq_vehicle_type'].astype(str).str.upper()

            # st.session_state['rfq_odvtmt_output_df_session'] = rfq_odvtmt_output_df
            st.session_state['rfq_odvt_output_df_session'] = rfq_odvt_output_df
            st.session_state.summary_stats = summary

            print('total rows in input df :', df_input_cluster.shape)
            # print('total rows in output df odvtmt level :', rfq_odvtmt_output_df.shape)
            print('total rows in output df odvt level :', rfq_odvt_output_df.shape)
        # from io import BytesIO

            progress_bar.progress(100)

            progress_text.text("✅ Processing Complete!")


    def to_excel_with_sheets(df1):
        highlight_columns = ['origin',	'destination', 'origin_state', 'destination_state',	'vehicle_type',	'load_mt','final_origin', 'final_destination', 'final_VT', 'final_tonnage']
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        file_name = 'rfq_pricing_processed_data.xlsx'

        # Step 1: Write to Excel
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df1.to_excel(writer, index=False, sheet_name='ODVTMT Output')
            # df2.to_excel(writer, index=False, sheet_name='ODVT Output')

        # Step 2: Apply formatting
        wb = load_workbook(file_name)

        def highlight_headers(ws, df):
            for col_name in highlight_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # openpyxl uses 1-based indexing
                    col_letter = get_column_letter(col_idx)
                    cell = ws[f"{col_letter}1"]  # Header is always row 1
                    cell.fill = header_fill
                    cell.font = header_font

        highlight_headers(wb['ODVTMT Output'], df1)
        # highlight_headers(wb['ODVT Output'], df2)

        wb.save(file_name)

        # Step 3: Return file as bytes
        with open(file_name, 'rb') as f:
            return f.read()


     

    # rfq_odvtmt_output_df1 = st.session_state.get('rfq_odvtmt_output_df_session')
    rfq_odvt_output_df1 = st.session_state.get('rfq_odvt_output_df_session')

  
    if rfq_odvt_output_df1 is not None:

        summary = st.session_state.summary_stats

        st.write('### RFQs Mapping Summary')
        st.write('Total Lanes Pricing request', summary['total_rfqs'])
        st.write('Total Lanes matched on origin & destination city (ODVT level)',  summary['OCDVT Level Matched'])
        st.write('Total Lanes matched on origin cluster & destination city (OCDVT level)',  summary['OCDVT Level Matched'])
        st.write('Total Lanes matched on origin & destination cluster (OCDCVT level)',  summary['OCDCVT Level matched'])
        

        st.write('### RFQs Mapped Pricing data')
        st.write(rfq_odvt_output_df1)

        excel_buffer = to_excel_with_sheets(rfq_odvt_output_df1)

        st.download_button(
            label="Download Processed Output",
            data=excel_buffer,
            file_name='rfq_pricing_processed_data.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        

